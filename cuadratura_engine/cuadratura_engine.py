import io
import re
import numpy as np
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOLERANCIA = 200

# Mapeos estáticos (Reglas de negocio v3)
PR_IDX = {
    'PR_AFP': 27, 'PR_SIS': 28, 'PR_CTA_AHORRO': 29,
    'PR_APVI': 42, 'PR_DEP_CONV': 43, 'PR_APVC_T': 47, 'PR_APVC_E': 48,
    'PR_FONASA': 69, 'PR_ISAPRE': 79, 'PR_ISAPRE_ADC': 80,
    'PR_CCAF_CRED': 84, 'PR_CCAF_DENTAL': 85, 'PR_CCAF_LEASING': 86, 'PR_CCAF_SEGVIDA': 87,
    'PR_CCAF_SALUD': 89, 'PR_SS': 93, 'PR_MUTUAL': 97, 'PR_AFC_T': 100, 'PR_AFC_E': 101,
    'PR_ASIG_FAM': 21, 'PR_ASIG_RETRO': 22, 'PR_REINTEGRO': 23,
}

LE_COD = {
    'LE_AFP': '3141', 'LE_SALUD7': '3143', 'LE_SALUD_ADC': '3144',
    'LE_AFC_T': '3151', 'LE_APVI_A': '3155', 'LE_APVI_B': '3156',
    'LE_APVC_A': '3157', 'LE_APVC_B': '3158', 'LE_CCAF': '3110',
    'LE_ASIG_FAM': '2311', 'LE_AFC_E': '4151', 'LE_MUTUAL': '4152', 'LE_4155': '4155',
}

CE_MAP = {
    'CE_CARGAS': 'CARGAS FAMILIARES', 'CE_CARGAS_RETRO': 'CARGAS RETROACTIVAS',
    'CE_AFC_E': 'SEGURO CESANTIA', 'CE_SIS': 'SIS',
    'CE_MUTUAL': 'SEGURO ACCIDENTES DEL TRABAJO',
    'CE_CAP': 'CAP. INDIV. AFP', 'CE_SS': 'SEGURO SOCIAL',
}

C = {'hdr': '1A3A5C', 'le': 'D6EAF8', 'pr': 'D5F5E3', 'ce': 'FADBD8',
     'ok': 'EAFAF1', 'warn': 'FDEDEC', 'alt': 'F8F9FA', 'sub': 'D5D8DC'}

def clean_rut(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    r = str(v).strip().replace('.', '').replace(' ', '').split('-')[0]
    r = re.sub(r'[^0-9]', '', r)
    return r if r and len(r) >= 6 else None

def num(s):
    return pd.to_numeric(s, errors='coerce').fillna(0.0)

# --- Parsers del Motor Core ---

def leer_lre_stream(file_stream):
    # Detectar codificación dinámicamente sobre el stream de bytes
    content = file_stream.read()
    file_stream.seek(0)
    enc = 'utf-8-sig'
    for e in ('utf-8-sig', 'latin-1', 'cp1252'):
        try:
            content.decode(e)
            enc = e
            break
        except UnicodeDecodeError:
            continue
            
    # Detectar separador
    sample = content.decode(enc, errors='ignore')[:2048]
    sep = ';' if sample.count(';') > 5 else ','
    
    df = pd.read_csv(io.BytesIO(content), sep=sep, dtype=str, encoding=enc)
    out = pd.DataFrame()
    out['RUT'] = df.iloc[:, 0].map(clean_rut)
    for key, cod in LE_COD.items():
        col = next((c for c in df.columns if f'({cod})' in c), None)
        out[key] = num(df[col]) if col else 0.0
    out = out[out['RUT'].notna()]
    out['LE_APV'] = out['LE_APVI_A'] + out['LE_APVI_B'] + out['LE_APVC_A'] + out['LE_APVC_B']
    return out.groupby('RUT', as_index=False).sum(numeric_only=True)

def leer_previred_stream(file_stream):
    df = pd.read_excel(file_stream, header=0, dtype=str)
    nombres = (df.iloc[:, 2].fillna('') + ' ' + df.iloc[:, 3].fillna('')
               + ' ' + df.iloc[:, 4].fillna('')).str.strip()
    
    # Extraer período desde la columna 8 (índice base 0 de Previred)
    periodo_txt = "Período"
    if len(df) > 0 and len(df.columns) > 8:
        val_periodo = str(df.iloc[0, 8]).strip()
        if re.match(r'^\d{6}$', val_periodo): # MMAAAA
            periodo_txt = f"{val_periodo[:2]}/{val_periodo[2:]}"
            
    out = pd.DataFrame({'RUT': df.iloc[:, 0].map(clean_rut), 'NOMBRE': nombres})
    for key, idx in PR_IDX.items():
        out[key] = num(df.iloc[:, idx])
    out = out[out['RUT'].notna()]
    agg = {k: 'sum' for k in PR_IDX}
    agg['NOMBRE'] = 'first'
    g = out.groupby('RUT', as_index=False).agg(agg)
    g['PR_SALUD7'] = g['PR_FONASA'] + g['PR_ISAPRE'] + g['PR_CCAF_SALUD']
    g['PR_APV'] = g['PR_APVI'] + g['PR_DEP_CONV'] + g['PR_APVC_T'] + g['PR_APVC_E']
    g['PR_CCAF'] = (g['PR_CCAF_CRED'] + g['PR_CCAF_DENTAL']
                    + g['PR_CCAF_LEASING'] + g['PR_CCAF_SEGVIDA'])
    g['PR_ASIG'] = g['PR_ASIG_FAM'] + g['PR_ASIG_RETRO'] + g['PR_REINTEGRO']
    return g, periodo_txt

def leer_ce_stream(file_stream):
    raw = pd.read_excel(file_stream, header=None, dtype=str)
    hdr_row = next(i for i in range(10)
                   if raw.iloc[i].astype(str).str.contains('R.U.T', na=False).any())
    df = pd.read_excel(file_stream, header=hdr_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    out = pd.DataFrame({'RUT': df['R.U.T.'].map(clean_rut)})
    for key, name in CE_MAP.items():
        col = next((c for c in df.columns if c.upper() == name), None)
        out[key] = num(df[col]) if col else 0.0
    out = out[out['RUT'].notna()]
    out['CE_CARGAS_T'] = out['CE_CARGAS'] + out['CE_CARGAS_RETRO']
    return out.groupby('RUT', as_index=False).sum(numeric_only=True)

# --- Estilos openpyxl ---
def _b():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _h(c, bg=None, fc='FFFFFF', sz=9):
    c.font = Font(bold=True, color=fc, name='Arial', size=sz)
    c.fill = PatternFill('solid', fgColor=bg or C['hdr'])
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _b()

def _n(c, v, bg=None, bold=False):
    c.value = v
    c.number_format = '#,##0;[Red](#,##0);"-"'
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = _b()
    c.font = Font(name='Arial', size=9, bold=bold)
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)

def _t(c, v, bg=None, bold=False, align='left', sz=9, italic=False, fc=None):
    c.value = v
    c.font = Font(bold=bold, name='Arial', size=sz, italic=italic, color=fc or '000000')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = _b()
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)

# --- Generadores de Planillas ---
CONCEPTOS = [
    ('AFP',              'LE_AFP',       'PR_AFP'),
    ('SIS + Seg.Social', 'LE_4155',      None),
    ('Salud 7%',         'LE_SALUD7',    'PR_SALUD7'),
    ('Salud Adicional',  'LE_SALUD_ADC', 'PR_ISAPRE_ADC'),
    ('AFC Trabajador',   'LE_AFC_T',     'PR_AFC_T'),
    ('AFC Empleador',    'LE_AFC_E',     'PR_AFC_E'),
    ('Mutual',           'LE_MUTUAL',    'PR_MUTUAL'),
    ('APV',              'LE_APV',       'PR_APV'),
    ('CCAF Crédito',     'LE_CCAF',      'PR_CCAF'),
    ('Asig. Familiar',   'LE_ASIG_FAM',  'PR_ASIG'),
]
CE_COLS = ['CE_AFC_E', 'CE_SIS', 'CE_MUTUAL', 'CE_SS', 'CE_CAP', 'CE_CARGAS_T']
CE_LBL = ['CE Ces. Empl.', 'CE SIS', 'CE Mutual', 'CE Seg.Social', 'CE Cap.Indiv.', 'CE Cargas Fam.']

def construir_hojas_excel(wb, df, periodo):
    # 1. Crear Hoja Comparación Detallada
    ws_comp = wb.create_sheet('Comparacion LE-PR')
    ws_comp.sheet_view.showGridLines = False
    n = len(df)
    last = n + 1

    headers = [('RUT', 13), ('Nombre', 26), ('Estado', 9)]
    col_of = {}
    ci = 4
    for lbl, cle, cpr in CONCEPTOS:
        if cpr is None:
            for sub, key in [('LRE 4155', 'LE_4155'), ('PR SIS', 'PR_SIS'), ('PR Seg.Soc.', 'PR_SS'), ('Dif', 'DIF_SIS')]:
                col_of[key] = ci
                headers.append((sub, 12)); ci += 1
        else:
            for pre, key in [('LRE', cle), ('PR', cpr), ('Dif', f'DIF_{cle}')]:
                col_of[key] = ci
                headers.append((pre if pre != 'Dif' else 'Dif', 12)); ci += 1
    for k, lbl in zip(CE_COLS, CE_LBL):
        col_of[k] = ci
        headers.append((lbl, 12)); ci += 1
    col_of['PR_CTA_AHORRO'] = ci
    headers.append(('PR Cta.Ahorro', 12)); ci += 1

    ws_comp.row_dimensions[1].height = 34
    for j, (lbl, w) in enumerate(headers, 1):
        bg = C['hdr']
        if lbl.startswith('LRE'): bg = '2874A6'
        elif lbl.startswith('PR'): bg = '1E8449'
        elif lbl.startswith('CE'): bg = '922B21'
        _h(ws_comp.cell(1, j), bg=bg)
        ws_comp.cell(1, j).value = lbl
        ws_comp.column_dimensions[get_column_letter(j)].width = w

    dif_letters = []
    for ri, (_, r) in enumerate(df.iterrows(), 2):
        _t(ws_comp.cell(ri, 1), r['RUT'])
        _t(ws_comp.cell(ri, 2), r.get('NOMBRE', ''))
        for lbl, cle, cpr in CONCEPTOS:
            if cpr is None:
                a, b, c, d = (get_column_letter(col_of[k]) for k in ('LE_4155', 'PR_SIS', 'PR_SS', 'DIF_SIS'))
                _n(ws_comp.cell(ri, col_of['LE_4155']), r['LE_4155'], bg=C['le'])
                _n(ws_comp.cell(ri, col_of['PR_SIS']), r['PR_SIS'], bg=C['pr'])
                _n(ws_comp.cell(ri, col_of['PR_SS']), r['PR_SS'], bg=C['pr'])
                ws_comp.cell(ri, col_of['DIF_SIS']).value = f'={a}{ri}-{b}{ri}-{c}{ri}'
                _h_diff_fmt(ws_comp.cell(ri, col_of['DIF_SIS']))
            else:
                a = get_column_letter(col_of[cle]); b = get_column_letter(col_of[cpr])
                _n(ws_comp.cell(ri, col_of[cle]), r[cle], bg=C['le'])
                _n(ws_comp.cell(ri, col_of[cpr]), r[cpr], bg=C['pr'])
                cell = ws_comp.cell(ri, col_of[f'DIF_{cle}'])
                cell.value = f'={a}{ri}-{b}{ri}'
                _h_diff_fmt(cell)
        for k in CE_COLS:
            _n(ws_comp.cell(ri, col_of[k]), r[k], bg=C['ce'])
        _n(ws_comp.cell(ri, col_of['PR_CTA_AHORRO']), r['PR_CTA_AHORRO'], bg=C['pr'])

        if ri == 2:
            dif_letters = [get_column_letter(col_of['DIF_SIS'])] + [get_column_letter(col_of[f'DIF_{cle}']) for _, cle, cpr in CONCEPTOS if cpr is not None]
        cond = ','.join(f'ABS({L}{ri})' for L in dif_letters)
        ws_comp.cell(ri, 3).value = f'=IF(MAX({cond})<={TOLERANCIA},"OK","DIFF")'
        ws_comp.cell(ri, 3).font = Font(bold=True, name='Arial', size=9)
        ws_comp.cell(ri, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws_comp.cell(ri, 3).border = _b()

    rt = last + 1
    _t(ws_comp.cell(rt, 1), 'TOTALES', bg=C['sub'], bold=True)
    _t(ws_comp.cell(rt, 2), '', bg=C['sub']); _t(ws_comp.cell(rt, 3), '', bg=C['sub'])
    for j in range(4, len(headers) + 1):
        L = get_column_letter(j)
        _n(ws_comp.cell(rt, j), f'=SUM({L}2:{L}{last})', bg=C['sub'], bold=True)
    ws_comp.freeze_panes = 'D2'

    # 2. Crear Cuadro "Resumen Imposiciones" (Fórmulas Vivas)
    ws_res = wb.create_sheet('Resumen Imposiciones', 0)
    ws_res.sheet_view.showGridLines = False
    S = "'Comparacion LE-PR'"
    def rng(key): return f"SUM({S}!{get_column_letter(col_of[key])}2:{get_column_letter(col_of[key])}{last})"

    ws_res.merge_cells('A1:G1')
    _h(ws_res['A1'], sz=12)
    ws_res['A1'] = f'DETALLE IMPOSICIONES POR PAGAR — {periodo}'
    ws_res.row_dimensions[1].height = 26

    cab = ['Concepto', 'Costo Empresa (KAME)', 'Libro LRE', 'Previred', 'Dif LRE−PR', 'Dif CE−PR', 'Nota']
    for j, t in enumerate(cab, 1): _h(ws_res.cell(2, j), bg='2E4057')
    for col, w in zip('ABCDEFG', [44, 18, 16, 16, 13, 13, 58]): ws_res.column_dimensions[col].width = w

    rows = [
        ('Imposiciones por Pagar (Seguro Cesantía Empleador)', f"={rng('CE_AFC_E')}", f"={rng('LE_AFC_E')}", f"={rng('PR_AFC_E')}", ''),
        ('Imposiciones por Pagar (SIS)', f"={rng('CE_SIS')}", f"={rng('LE_4155')}-{rng('PR_SS')}", f"={rng('PR_SIS')}", 'LRE cód. 4155 trae SIS + Seguro Social juntos; se descuenta el SS para comparar.'),
        ('Imposiciones por Pagar (Mutual)', f"={rng('CE_MUTUAL')}", f"={rng('LE_MUTUAL')}", f"={rng('PR_MUTUAL')}", ''),
        ('Imposiciones por Pagar (SS Expectativa de Vida)', f"={rng('CE_SS')}", f"={rng('PR_SS')}", f"={rng('PR_SS')}", 'En el LRE va incluido dentro del cód. 4155. En Previred: campo 94 "Bonos Gobierno (CCAF)".'),
        ('Imposiciones por Pagar (Aporte Cap. Individual AFP)', f"={rng('CE_CAP')}", 0, 0, 'Solo Costo Empresa lo separa; en LRE y Previred va dentro de la cotización AFP.'),
        ('Imposiciones por Pagar (AFP)', '=C8-B7', f"={rng('LE_AFP')}", f"={rng('PR_AFP')}", 'KAME derivado: LRE AFP menos Aporte Cap. Individual.'),
        ('Imposiciones por Pagar (Salud)', None, f"={rng('LE_SALUD7')}", f"={rng('PR_SALUD7')}", 'PR = Fonasa + Isapre + cotización 7% vía CCAF.'),
        ('Imposiciones por Pagar (Adicional Salud)', None, f"={rng('LE_SALUD_ADC')}", f"={rng('PR_ISAPRE_ADC')}", ''),
        ('Imposiciones por Pagar (Seguro Cesantía)', None, f"={rng('LE_AFC_T')}", f"={rng('PR_AFC_T')}", 'Aporte del trabajador.'),
        ('Imposiciones por Pagar (APV)', None, f"={rng('LE_APV')}", f"={rng('PR_APV')}", ''),
        ('Imposiciones por Pagar (cuenta de ahorro)', None, None, f"={rng('PR_CTA_AHORRO')}", 'El LRE no tiene código separado para la cta. 2.'),
    ]

    r = 3
    first_data = r
    for concepto, fce, flre, fpr, nota in rows:
        _t(ws_res.cell(r, 1), concepto)
        if fce is not None: _n(ws_res.cell(r, 2), fce, bg='C6E0B4')
        else: _t(ws_res.cell(r, 2), '—', align='center', fc='999999')
        if flre is not None: _n(ws_res.cell(r, 3), flre)
        else: _t(ws_res.cell(r, 3), '—', align='center', fc='999999')
        _n(ws_res.cell(r, 4), fpr)
        ws_res.cell(r, 5).value = f'=IF(C{r}="—","",C{r}-D{r})' if flre is None else f'=C{r}-D{r}'
        _h_diff_fmt(ws_res.cell(r, 5))
        ws_res.cell(r, 6).value = f'=B{r}-D{r}' if fce is not None else ''
        _h_diff_fmt(ws_res.cell(r, 6))
        _t(ws_res.cell(r, 7), nota, italic=True, fc='595959', sz=8)
        r += 1
    last_data = r - 1

    _t(ws_res.cell(r, 1), 'Sub total imposiciones por pagar', bold=True, bg=C['sub'])
    for col in 'BCD':
        c = ws_res.cell(r, 'BCD'.index(col) + 2)
        c.value = f'=SUM({col}{first_data}:{col}{last_data})'
        _n(c, c.value, bg=C['sub'], bold=True)
    ws_res.cell(r, 5).value = f'=SUM(E{first_data}:E{last_data})'
    _h_diff_fmt(ws_res.cell(r, 5)); ws_res.cell(r, 5).font = Font(bold=True, name='Arial', size=9); ws_res.cell(r, 5).fill = PatternFill('solid', fgColor=C['sub'])
    sub_row = r
    r += 2

    _t(ws_res.cell(r, 1), 'CCAF (crédito social y otros descuentos CCAF)')
    _t(ws_res.cell(r, 2), '—', align='center', fc='999999')
    _n(ws_res.cell(r, 3), f"={rng('LE_CCAF')}")
    _n(ws_res.cell(r, 4), f"={rng('PR_CCAF')}")
    ws_res.cell(r, 5).value = f'=C{r}-D{r}'; _h_diff_fmt(ws_res.cell(r, 5))
    _t(ws_res.cell(r, 7), 'PR = créditos + dental + leasing + seguro de vida CCAF.', italic=True, fc='595959', sz=8)
    ccaf_row = r
    r += 1

    _t(ws_res.cell(r, 1), 'Asignación familiar (a descontar)')
    _n(ws_res.cell(r, 2), f"={rng('CE_CARGAS_T')}", bg='C6E0B4')
    ws_res.cell(r, 3).value = f"=-{rng('LE_ASIG_FAM')}"; _h_diff_fmt(ws_res.cell(r, 3))
    ws_res.cell(r, 4).value = f"=-{rng('PR_ASIG')}"; _h_diff_fmt(ws_res.cell(r, 4))
    ws_res.cell(r, 5).value = f'=C{r}-D{r}'; _h_diff_fmt(ws_res.cell(r, 5))
    ws_res.cell(r, 6).value = f'=B{r}-D{r}'; _h_diff_fmt(ws_res.cell(r, 6))
    _t(ws_res.cell(r, 7), 'Crédito fiscal: reduce el total a pagar.', italic=True, fc='595959', sz=8)
    asig_row = r
    r += 1

    _t(ws_res.cell(r, 1), 'Total imposiciones a pagar', bold=True, bg=C['sub'])
    for col in 'BCD':
        j = 'BCD'.index(col) + 2
        ws_res.cell(r, j).value = f'={col}{sub_row}+{col}{ccaf_row}+{col}{asig_row}' if col != 'B' else f'=B{sub_row}+B{asig_row}'
        _n(ws_res.cell(r, j), ws_res.cell(r, j).value, bg=C['sub'], bold=True)
    ws_res.cell(r, 5).value = f'=E{sub_row}+E{ccaf_row}+E{asig_row}'
    _h_diff_fmt(ws_res.cell(r, 5)); ws_res.cell(r, 5).font = Font(bold=True, name='Arial', size=9); ws_res.cell(r, 5).fill = PatternFill('solid', fgColor=C['sub'])

    # 3. Hoja KPIs Alterna
    ws_kpi = wb.create_sheet('KPIs Cuadratura', 1)
    ws_kpi.sheet_view.showGridLines = False
    ws_kpi.merge_cells('A1:C1')
    _h(ws_kpi['A1'], sz=11)
    ws_kpi['A1'] = f'RESUMEN GENERAL CUADRATURA — {periodo}'
    
    datos_kpi = [
        ('Trabajadores analizados', n, 'Universo único de RUTs.'),
        ('Trabajadores OK', f'=COUNTIF({S}!C2:C{last},"OK")', 'Sin diferencias significativas.'),
        ('Trabajadores con diferencias', f'=COUNTIF({S}!C2:C{last},"DIFF")', 'Requieren revisión en la pestaña detallada.'),
        ('% OK', f'=B3/B2', 'Ratio de coincidencia.'),
    ]
    for idx, (lbl, formula_val, n_desc) in enumerate(datos_kpi, 2):
        _t(ws_kpi.cell(idx, 1), lbl, bold=True)
        ws_kpi.cell(idx, 2).value = formula_val
        if lbl == '% OK': ws_kpi.cell(idx, 2).number_format = '0.0%'
        _t(ws_kpi.cell(idx, 3), n_desc, italic=True, fc='595959')
    ws_kpi.column_dimensions['A'].width = 30
    ws_kpi.column_dimensions['B'].width = 14
    ws_kpi.column_dimensions['C'].width = 62

def _h_diff_fmt(c):
    c.number_format = '#,##0;[Red](#,##0);"-"'
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = _b()
    c.font = Font(name='Arial', size=9)

# --- Orquestador Maestro ---
def ejecutar_proceso_cuadratura(lre_file, pr_file, ce_file):
    df_le = leer_lre_stream(lre_file)
    df_pr, periodo = leer_previred_stream(pr_file)
    df_ce = leer_ce_stream(ce_file)
    
    # Cruce maestro por RUT (Regla v3: no elimina líneas de movimientos de personal, las consolida)
    df = pd.merge(df_pr, df_le, on='RUT', how='outer')
    df = pd.merge(df, df_ce, on='RUT', how='outer')
    
    num_cols = df.select_dtypes(include='number').columns
    df[num_cols] = df[num_cols].fillna(0)
    df['NOMBRE'] = df['NOMBRE'].fillna('')
    df = df.sort_values('NOMBRE').reset_index(drop=True)
    
    wb = Workbook()
    wb.remove(wb.active) # Eliminar hoja por defecto
    construir_hojas_excel(wb, df, periodo)
    
    # Guardar en buffer en lugar de disco rígido
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, periodo, len(df)