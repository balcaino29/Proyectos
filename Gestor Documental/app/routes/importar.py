import io
from datetime import date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
import openpyxl

from app.database import get_db
from app import models

router = APIRouter(prefix="/api", tags=["importar"])

ESTADO_COL = 18   # columna R (1-indexado): ESTATUS
HOJA_NOMBRE = "MASTER INGENIER\u00cdA CLIENTE"


def safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s not in ("-", "N/A", "n/a") else None


def safe_date(v):
    if isinstance(v, date):
        return v
    return None


def safe_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


@router.post("/importar")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    # Buscar la hoja (puede tener caracteres especiales)
    ws = None
    for name in wb.sheetnames:
        if "MASTER" in name.upper() and "ING" in name.upper() and "CLIENTE" in name.upper():
            ws = wb[name]
            break
    if ws is None:
        raise HTTPException(status_code=400, detail="No se encontró la hoja MASTER INGENIERÍA CLIENTE")

    docs_creados = 0
    emisiones_creadas = 0

    # Filas de datos desde fila 4 (índice 4 en 1-based)
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Columna B (índice 1) = item número
        item_val = row[1]
        codigo = safe_str(row[4])  # E: CÓDIGO SEGURA
        descripcion = safe_str(row[7])  # H: DESCRIPCIÓN

        # Saltar filas sin código ni descripción
        if not codigo and not descripcion:
            continue

        doc = models.Documento(
            item=safe_int(item_val),
            tipo_ing=safe_str(row[2]),
            sala_tag=safe_str(row[3]),
            codigo_segura=codigo,
            codigo_inn=safe_str(row[5]),
            codigo_codelco=safe_str(row[6]),
            descripcion=descripcion,
            compromiso_entrega=safe_str(row[8]),
            criticidad=safe_str(row[9]),
            deadline_cliente=safe_str(row[10]),
            comentario_interno=safe_str(row[11]),
            disciplina=safe_str(row[12]),
            revision_actual=str(row[13]) if row[13] is not None else None,
            acta_trl=safe_int(row[14]),
            fecha_ultimo_envio=safe_date(row[15]),
            fecha_respuesta_cliente=safe_date(row[16]),
            estatus=safe_str(row[17]),
        )
        db.add(doc)
        db.flush()
        docs_creados += 1

        # Parsear emisiones: columnas a partir de índice 21 (V)
        # Grupos de 4 columnas: REVISION, TRANSMITTAL_SEH, TIEMPO_RESPUESTA, FECHA_ENVIO,
        # TRANSMITTAL_CLIENTE, FECHA_RESPUESTA, TIEMPO_RESP_CLIENTE, OBS
        # Desde la fila: col 21=V (revisión 1era emisión)
        # Patrón por emisión (8 columnas por grupo):
        # [rev, trl_seh, tiempo, fecha_envio, trl_cliente, fecha_resp, tiempo_resp, obs]
        col_start = 21  # índice base 0 → columna V
        emision_num = 1
        while col_start + 7 < len(row):
            rev = safe_str(row[col_start])
            trl_seh = safe_int(row[col_start + 1])
            fecha_envio = safe_date(row[col_start + 3])
            trl_cliente = safe_str(row[col_start + 4])
            fecha_resp = safe_date(row[col_start + 5])
            tiempo_resp = safe_int(row[col_start + 6])
            obs = safe_str(row[col_start + 7])

            if fecha_envio or trl_seh:
                emision = models.Emision(
                    documento_id=doc.id,
                    numero=emision_num,
                    revision=rev,
                    transmittal_seh=trl_seh,
                    fecha_envio=fecha_envio,
                    transmittal_cliente=trl_cliente,
                    fecha_respuesta=fecha_resp,
                    tiempo_respuesta=tiempo_resp,
                    observaciones=obs,
                )
                db.add(emision)
                emisiones_creadas += 1

            emision_num += 1
            col_start += 8

    db.commit()
    return {"importados": docs_creados, "emisiones": emisiones_creadas}
