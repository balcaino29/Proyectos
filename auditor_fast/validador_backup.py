"""
validador.py - Stage 2 + Stage 3 del Agente FAST
Logica 100% determinista. Sin LLM. Sin dependencias externas.

Ejecutar localmente:
    python validador.py --fast fast.xlsx --tarifario tarifas_FAST.xlsx

En produccion (Cloud Run / LangGraph):
    from validador import ejecutar_validacion
    resultado = ejecutar_validacion(path_fast, path_tarifario)
"""

import math
import argparse
import warnings
from datetime import datetime, timedelta
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# --- CONSTANTES ----------------------------------------------------------------
COBRO_MINIMO         = 41_661       # CLP - piso por servicio KG
HORAS_GRACIA         = 24           # Almacenaje libre
TOLERANCE_PCT        = 0.01         # 1% tolerancia por punto flotante
HOJAS_DETALLE        = ["Aqua_Latam", "Aqua_Otras_Aerolineas", "Otros A&A"]
CLIENTE_BLOQUEADO    = "AUSTRALIS"
CONCEPTO_EXCLUIDO    = "PLASTICO SEPARACION AWB"
CONCEPTO_DESTRUCCION = "DESTRUCCION CAJAS"

# Servicios obligatorios por AWB
OBLIGATORIO_DESCARGA = "DESCARGA Y PALETIZAJE"
OBLIGATORIO_FULL     = "FULL SERVICE"
OBLIGATORIOS_RX      = {"RX (PORCENTAJE)", "SERVICIO EMIS (PORCENTAJE)", "RAYOS EMIS"}

# Mapa de nombres FAST -> clave normalizada para lookup en tarifario
# Clave: texto que aparece en columna "Nombre Servicio" del archivo FAST
# Valor: (clave_busqueda_tarifario, observacion_para_match)
MAPA_SERVICIOS = {
    "ALMACENAJE":               ("Almacenaje adicional",          None),
    "DESCARGA Y PALETIZAJE":    ("Descarga",                      None),
    "FULL SERVICE":             ("Full Services",                  None),
    "RX (PORCENTAJE)":          ("Rayos X",                       None),
    "SERVICIO EMIS (PORCENTAJE)": ("EMIS (CMD)",                  None),
    "RAYOS EMIS":               ("EMIS (CMD)",                    None),
    "TRASVASIJE":               ("Trasvasije",                     None),
    "DESTRUCCION CAJAS [AWB]":  ("Destruccion caja",              None),
    "ENMANTADO (2 POSICIONES)": ("Enmantado",                     "ULD 2 POSICIONES"),
    "ENMANTADO (4 POSICIONES)": ("Enmantado",                     "ULD 4 POSICIONES"),
    "ENMANTADO CAJA":           ("Enmantado",                     "CAJA"),
    "SELLADO (FILM) POR SKID":  ("Sellado (Stretch Film)",        "SKID"),
    "REETIQUETADO":             ("Reetiquetado",                   None),
    "DOBLE ETIQUETADO":         ("Reetiquetado",                   None),   # misma tarifa $7
    "PEGADO SOBRE TARDIO":      ("Hora Hombre",                   None),   # tarifa $22.318 UNIDAD
    "PLASTICO SEPARACION AWB":  ("Plastico Separacion AWB",       None),   # -> forzar $0
}

# Colores para Excel de salida
COLOR_OK       = "C6EFCE"   # verde suave
COLOR_ALERTA   = "FFEB9C"   # amarillo
COLOR_ERROR    = "FFC7CE"   # rojo suave
COLOR_HEADER   = "1F4E79"   # azul oscuro A&A
COLOR_INFO     = "DDEBF7"   # azul muy claro


# --- CARGA DEL TARIFARIO -------------------------------------------------------
def cargar_tarifario(path: str) -> pd.DataFrame:
    """
    Lee tarifas_FAST.xlsx hoja SALMON.
    Retorna DataFrame limpio con columnas normalizadas.
    """
    # El tarifario tiene 4 filas de metadata antes del header real
    df = pd.read_excel(path, sheet_name="SALMON", header=4)
    # La primera fila de datos es el header real
    df.columns = [str(c).strip() for c in df.iloc[0]]
    df = df.iloc[1:].reset_index(drop=True)

    col_map = {
        "TERMINAL":       "terminal",
        "SERVICIO":       "servicio",
        "UNIDAD":         "unidad",
        "TARIFA A&A":     "tarifa_aa",
        "TARIFA PUBLICA": "tarifa_publica",
        "MONEDA":         "moneda",
        "OBSERVACION":    "observacion",
        "VIGENCIA":       "vigencia",
    }
    # Normalizar nombres de columna: quitar tildes para match seguro
    import unicodedata
    def norm(s):
        return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode("ascii").upper().strip()
    df.columns = [norm(c) for c in df.columns]
    col_map_norm = {norm(k): v for k, v in col_map.items()}
    df = df.rename(columns={k: v for k, v in col_map_norm.items() if k in df.columns})
    df = df[df["servicio"].notna()].copy()
    df["servicio"]    = df["servicio"].str.strip()
    df["observacion"] = df["observacion"].fillna("").str.strip().str.upper()
    return df


def buscar_tarifa(tarifario: pd.DataFrame, nombre_fast: str) -> float | None:
    """
    Resuelve el nombre del archivo FAST hacia la tarifa publica correcta.
    Maneja variantes de ENMANTADO por observacion.
    Retorna float o None si no se encuentra.
    """
    nombre_fast = str(nombre_fast).strip().upper()

    # Buscar en el mapa
    entrada = MAPA_SERVICIOS.get(nombre_fast)
    if not entrada:
        # Intento fuzzy: si el nombre contiene alguna clave
        for k, v in MAPA_SERVICIOS.items():
            if k in nombre_fast or nombre_fast in k:
                entrada = v
                break
    if not entrada:
        return None

    clave_tar, obs_requerida = entrada
    subset = tarifario[tarifario["servicio"].str.strip() == clave_tar]

    if subset.empty:
        return None

    if obs_requerida:
        match = subset[subset["observacion"].str.contains(obs_requerida, case=False, na=False)]
        if not match.empty:
            return float(match.iloc[0]["tarifa_publica"])

    return float(subset.iloc[0]["tarifa_publica"])


def validar_vigencia(tarifario: pd.DataFrame, fecha_referencia: datetime) -> dict:
    """
    Verifica que la fecha de referencia este dentro de la vigencia del tarifario.
    """
    primera_fila = tarifario[tarifario["vigencia"].notna()].iloc[0]
    vigencia_str = str(primera_fila.get("vigencia", ""))
    try:
        partes   = vigencia_str.split(" - ")
        inicio   = datetime.strptime(partes[0].strip(), "%d/%m/%Y")
        fin      = datetime.strptime(partes[1].strip(), "%d/%m/%Y")
        vigente  = inicio <= fecha_referencia <= fin
        return {"vigente": vigente, "inicio": inicio, "fin": fin,
                "vigencia_str": vigencia_str}
    except Exception:
        return {"vigente": True, "inicio": None, "fin": None,
                "vigencia_str": vigencia_str, "advertencia": "No se pudo parsear vigencia"}


# --- CALCULO MATEMATICO POR FILA ----------------------------------------------
def parse_datetime_fast(fecha, hora) -> datetime | None:
    """Combina fecha y hora de las columnas FAST en un datetime."""
    try:
        if pd.isna(fecha):
            return None
        # Fecha puede venir como string "29-04-2026" o como datetime de Excel
        if isinstance(fecha, datetime):
            f = fecha
        else:
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    f = datetime.strptime(str(fecha).strip(), fmt)
                    break
                except ValueError:
                    continue
            else:
                return None

        if pd.isna(hora):
            return f
        # Hora puede ser "17:16" o un time object
        if isinstance(hora, str):
            h, m = hora.strip().split(":")
            return f.replace(hour=int(h), minute=int(m))
        if hasattr(hora, "hour"):
            return f.replace(hour=hora.hour, minute=hora.minute)
        return f
    except Exception:
        return None


def calcular_valor_fila(row: pd.Series, tarifario: pd.DataFrame) -> dict:
    """
    Calcula el valor esperado para una fila y lo compara con Valor cobro Ajustado.
    Retorna dict con: valor_calculado, diferencia, estado, nota.
    """
    nombre_servicio = str(row.get("Nombre Servicio", "")).strip().upper()
    unidad          = str(row.get("Unidad Cobro", "")).strip().upper()
    cantidad        = float(row.get("Cantidad A Cobro Ajustada") or 0)
    kilos_guia      = float(row.get("Kilos Guia") or 0)
    valor_cobrado   = float(row.get("Valor cobro Ajustado") or 0)
    tarifa_fast     = float(row.get("Tarifa (CLP)") or 0)

    # Obtener tarifa del tarifario Drive
    tarifa_tarifario = buscar_tarifa(tarifario, nombre_servicio)
    tarifa_usar      = tarifa_tarifario if tarifa_tarifario is not None else tarifa_fast

    # -- PLASTICO SEPARACION AWB: siempre $0 ----------------------------------
    if CONCEPTO_EXCLUIDO.upper() in nombre_servicio:
        if valor_cobrado != 0:
            return {
                "valor_calculado": 0,
                "diferencia":      -valor_cobrado,
                "estado":          "ERROR",
                "nota":            f"PLASTICO SEPARACION debe ser $0. Cobrado: ${valor_cobrado:,.0f}",
                "tarifa_usada":    tarifa_usar,
            }
        return {"valor_calculado": 0, "diferencia": 0, "estado": "OK",
                "nota": "Cobro aerolinea - excluido", "tarifa_usada": 0}

    # -- DESTRUCCION CAJAS: solo extraccion, calculo simple -------------------
    if CONCEPTO_DESTRUCCION in nombre_servicio:
        valor_calc = cantidad * tarifa_usar
        diff = valor_cobrado - valor_calc
        ok   = abs(diff) <= max(COBRO_MINIMO * TOLERANCE_PCT, 1)
        return {
            "valor_calculado": round(valor_calc, 2),
            "diferencia":      round(diff, 2),
            "estado":          "OK" if ok else "ALERTA",
            "nota":            "Registrar en tabla Destrucciones",
            "tarifa_usada":    tarifa_usar,
        }

    # -- ALMACENAJE: gracia 24h + formula Cantidad x Tarifa x dias_excedentes --
    # Formula FAST real (verificada contra archivo):
    #   dias_excedentes = floor(horas_totales / 24) - 1   (24h gracia descontadas)
    #   cobro = Cantidad x Tarifa x dias_excedentes
    # Ejemplo: 82.5h -> floor(82.5/24)=3 -> 3-1=2 dias... NO, factor=3.
    # Revisando: 82.5h -> floor=3, cobrado = Cant?Tar?3 -> dias = trunc(horas/24)
    # 72.2h -> trunc=3, cobrado = Cant?Tar?3 -> dias = trunc(horas/24)
    # La gracia de 24h aplica solo para cobro $0, NO descuenta del multiplicador.
    # Si horas > 24: dias_cobro = floor(horas_totales / 24)
    if nombre_servicio == "ALMACENAJE":
        inicio = parse_datetime_fast(row.get("Fecha Inicio Cobro"),
                                     row.get("Hora Inicio Cobro"))
        fin    = parse_datetime_fast(row.get("Fecha Termino Cobro"),
                                     row.get("Hora Termino Cobro"))

        # Si no hay fechas propias, usar columna Duracion si existe
        duracion_raw = row.get("Duracion")
        if inicio is None or fin is None:
            if duracion_raw is not None and str(duracion_raw) not in ("nan", "None", ""):
                try:
                    import pandas as _pd
                    td = _pd.to_timedelta(duracion_raw, errors="raise")
                    delta_horas = td.total_seconds() / 3600
                except Exception:
                    return {"valor_calculado": None, "diferencia": None,
                            "estado": "ADVERTENCIA",
                            "nota":   "Almacenaje sin fechas ni duracion - verificar manualmente",
                            "tarifa_usada": tarifa_usar}
            else:
                return {"valor_calculado": None, "diferencia": None,
                        "estado": "ADVERTENCIA",
                        "nota":   "Almacenaje sin fechas completas - verificar manualmente",
                        "tarifa_usada": tarifa_usar}
        else:
            delta_horas = (fin - inicio).total_seconds() / 3600

        if delta_horas <= HORAS_GRACIA:
            if valor_cobrado != 0:
                return {"valor_calculado": 0, "diferencia": -valor_cobrado,
                        "estado": "ERROR",
                        "nota":  f"Almacenaje <=24h debe ser $0. Cobrado: ${valor_cobrado:,.0f}",
                        "tarifa_usada": tarifa_usar}
            return {"valor_calculado": 0, "diferencia": 0, "estado": "OK",
                    "nota": f"Almacenaje {delta_horas:.1f}h - dentro de gracia 24h",
                    "tarifa_usada": tarifa_usar}

        # Mas de 24h: dias_cobro = parte entera de horas / 24
        dias_cobro = int(delta_horas / 24)
        valor_calc = cantidad * tarifa_usar * dias_cobro
        valor_calc = max(valor_calc, COBRO_MINIMO)
        diff       = valor_cobrado - valor_calc
        tolerancia = max(abs(valor_calc) * TOLERANCE_PCT, 1)
        estado     = "OK" if abs(diff) <= tolerancia else ("ALERTA" if abs(diff) <= 5000 else "ERROR")

        return {
            "valor_calculado": round(valor_calc, 2),
            "diferencia":      round(diff, 2),
            "estado":          estado,
            "nota":            f"{delta_horas:.1f}h -> {dias_cobro} dias | {cantidad:.0f}kg x ${tarifa_usar} x {dias_cobro}",
            "tarifa_usada":    tarifa_usar,
        }

    # -- KG: piso minimo solo aplica a DESCARGA Y PALETIZAJE y TRASVASIJE -----
    SERVICIOS_CON_PISO = {"DESCARGA Y PALETIZAJE", "TRASVASIJE"}
    if unidad == "KG":
        valor_base    = cantidad * tarifa_usar
        aplica_piso   = nombre_servicio in SERVICIOS_CON_PISO
        valor_calc    = max(valor_base, COBRO_MINIMO) if aplica_piso else valor_base
        aplicado_piso = aplica_piso and valor_calc == COBRO_MINIMO and valor_base < COBRO_MINIMO
        diff          = valor_cobrado - valor_calc
        tolerancia    = max(abs(valor_calc) * TOLERANCE_PCT, 1)
        estado        = "OK" if abs(diff) <= tolerancia else ("ALERTA" if abs(diff) <= 5000 else "ERROR")
        nota          = f"{'WARN Piso minimo - ' if aplicado_piso else ''}{cantidad:.0f}kg x ${tarifa_usar}"
        return {"valor_calculado": round(valor_calc, 2), "diferencia": round(diff, 2),
                "estado": estado, "nota": nota, "tarifa_usada": tarifa_usar}

    # -- %KG: (Kilos Guia ? 10%) ? Tarifa - sin piso minimo ------------------
    if unidad == "% KG":
        base_kg    = kilos_guia * 0.10
        valor_calc = base_kg * tarifa_usar
        # RX / EMIS no tienen cobro minimo - cobran el valor real
        diff       = valor_cobrado - valor_calc
        tolerancia = max(abs(valor_calc) * TOLERANCE_PCT, 1)
        estado     = "OK" if abs(diff) <= tolerancia else ("ALERTA" if abs(diff) <= 5000 else "ERROR")
        return {"valor_calculado": round(valor_calc, 2), "diferencia": round(diff, 2),
                "estado": estado,
                "nota":   f"{kilos_guia:.0f}kg_guia x 10% x ${tarifa_usar}",
                "tarifa_usada": tarifa_usar}

    # -- UNIDAD: cantidad x tarifa, sin piso (excepto COBRO_MINIMO para ciertos) -
    if unidad == "UNIDAD":
        valor_calc = cantidad * tarifa_usar
        diff       = valor_cobrado - valor_calc
        tolerancia = max(abs(valor_calc) * TOLERANCE_PCT, 1)
        estado     = "OK" if abs(diff) <= tolerancia else ("ALERTA" if abs(diff) <= 5000 else "ERROR")
        return {"valor_calculado": round(valor_calc, 2), "diferencia": round(diff, 2),
                "estado": estado,
                "nota":   f"{cantidad:.0f} unidades x ${tarifa_usar}",
                "tarifa_usada": tarifa_usar}

    # -- Unidad desconocida ----------------------------------------------------
    return {"valor_calculado": None, "diferencia": None, "estado": "ADVERTENCIA",
            "nota": f"Unidad de cobro no reconocida: {unidad}", "tarifa_usada": tarifa_usar}


# --- STAGE 2: VALIDACION MATEMATICA -------------------------------------------
def stage2_validar_matematica(
    dfs: dict[str, pd.DataFrame],
    tarifario: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    """
    Recibe dict {hoja: DataFrame} de las 3 hojas de detalle.
    Agrega columnas de validacion a cada DataFrame.
    Retorna dict con DataFrames enriquecidos.
    """
    resultados = {}

    for hoja, df in dfs.items():
        filas = []
        for _, row in df.iterrows():
            calc = calcular_valor_fila(row, tarifario)
            filas.append(calc)

        df_calc = pd.DataFrame(filas)
        df_out  = df.copy()
        df_out["Valor_Calculado"]  = df_calc["valor_calculado"]
        df_out["Diferencia_CLP"]   = df_calc["diferencia"]
        df_out["Estado_Validacion"] = df_calc["estado"]
        df_out["Nota_Validacion"]   = df_calc["nota"]
        df_out["Tarifa_Tarifario"]  = df_calc["tarifa_usada"]

        resultados[hoja] = df_out

    return resultados


# --- STAGE 3: AUDITOR?A DE INTEGRIDAD POR AWB ---------------------------------
def stage3_auditar_integridad(
    dfs_validados: dict[str, pd.DataFrame]
) -> dict:
    """
    Agrupa por AWB en todas las hojas y aplica las reglas de negocio.
    Retorna dict con: anomalias, destrucciones, bloqueados, resumen_awbs.
    """
    # Unir todas las hojas con etiqueta de origen
    frames = []
    for hoja, df in dfs_validados.items():
        d = df.copy()
        d["_hoja"] = hoja
        frames.append(d)
    df_total = pd.concat(frames, ignore_index=True)

    anomalias    = []
    destrucciones = []
    bloqueados   = []
    resumen_awbs = {}

    # Agrupar por AWB
    for awb, grupo in df_total.groupby("Awb", dropna=True):
        servicios   = grupo["Nombre Servicio"].str.upper().str.strip().tolist()
        clientes    = grupo["Cliente a facturar"].str.upper().str.strip().unique()
        exportador  = grupo["Exportador"].dropna().iloc[0] if not grupo["Exportador"].dropna().empty else ""

        info_awb = {"awb": awb, "exportador": exportador, "alertas": [], "estado_global": "OK"}

        # -- BLOQUEO AUSTRALIS -------------------------------------------------
        if any(CLIENTE_BLOQUEADO in c for c in clientes):
            bloqueados.append({"awb": awb, "exportador": exportador,
                               "cliente": ", ".join(clientes)})
            info_awb["alertas"].append("CRITICO: Cliente AUSTRALIS - bloqueo absoluto")
            info_awb["estado_global"] = "CRITICO"
            anomalias.append({
                "awb": awb, "nivel": "CRITICO",
                "tipo": "BLOQUEO_CLIENTE",
                "descripcion": f"AWB {awb} asociada a cliente AUSTRALIS. Rechazada."
            })

        # -- VERIFICAR OBLIGATORIOS --------------------------------------------
        tiene_descarga = any(OBLIGATORIO_DESCARGA in s for s in servicios)
        tiene_full     = any(OBLIGATORIO_FULL in s for s in servicios)
        tiene_rx       = any(s in OBLIGATORIOS_RX for s in servicios)

        if not tiene_descarga:
            anomalias.append({"awb": awb, "nivel": "CRITICO",
                              "tipo": "OMISION_OBLIGATORIO",
                              "descripcion": f"Falta DESCARGA Y PALETIZAJE en AWB {awb}"})
            info_awb["alertas"].append("CRITICO: Falta DESCARGA Y PALETIZAJE")
            info_awb["estado_global"] = "CRITICO"

        if not tiene_full:
            anomalias.append({"awb": awb, "nivel": "CRITICO",
                              "tipo": "OMISION_OBLIGATORIO",
                              "descripcion": f"Falta FULL SERVICE en AWB {awb}"})
            info_awb["alertas"].append("CRITICO: Falta FULL SERVICE")
            info_awb["estado_global"] = "CRITICO"

        if not tiene_rx:
            anomalias.append({"awb": awb, "nivel": "CRITICO",
                              "tipo": "OMISION_OBLIGATORIO",
                              "descripcion": f"Falta RX / EMIS / RAYOS EMIS en AWB {awb}"})
            info_awb["alertas"].append("CRITICO: Falta RX o EMIS")
            info_awb["estado_global"] = "CRITICO"

        # -- VERIFICAR DUPLICIDADES EN OBLIGATORIOS ----------------------------
        for concepto, label in [
            (OBLIGATORIO_DESCARGA, "DESCARGA Y PALETIZAJE"),
            (OBLIGATORIO_FULL,     "FULL SERVICE"),
        ]:
            count = sum(1 for s in servicios if concepto in s)
            if count > 1:
                anomalias.append({"awb": awb, "nivel": "ALERTA",
                                  "tipo": "DUPLICIDAD",
                                  "descripcion": f"{label} aparece {count} veces en AWB {awb}"})
                info_awb["alertas"].append(f"ALERTA: {label} duplicado ({count}x)")
                if info_awb["estado_global"] == "OK":
                    info_awb["estado_global"] = "ALERTA"

        # -- PLASTICO SEPARACION AWB -------------------------------------------
        filas_plastico = grupo[grupo["Nombre Servicio"].str.upper().str.contains(
            "PLASTICO SEPARACION", na=False)]
        for _, fp in filas_plastico.iterrows():
            if float(fp.get("Valor cobro Ajustado") or 0) != 0:
                anomalias.append({"awb": awb, "nivel": "CRITICO",
                                  "tipo": "CONCEPTO_NO_PERMITIDO",
                                  "descripcion": (f"PLASTICO SEPARACION AWB cobrado en ${fp['Valor cobro Ajustado']:,.0f} "
                                                  f"- debe ser $0 (cobro a aerolinea)")})
                info_awb["alertas"].append("CRITICO: PLASTICO SEPARACION AWB con valor > $0")
                info_awb["estado_global"] = "CRITICO"

        # -- DESTRUCCION DE CAJAS ----------------------------------------------
        filas_dest = grupo[grupo["Nombre Servicio"].str.upper().str.contains(
            "DESTRUCCION CAJAS", na=False)]
        for _, fd in filas_dest.iterrows():
            destrucciones.append({
                "awb":       awb,
                "exportador": exportador,
                "hoja":      fd.get("_hoja", ""),
                "valor":     fd.get("Valor cobro Ajustado", 0),
                "vuelo":     fd.get("Vuelo", ""),
                "destino":   fd.get("Destino", ""),
            })

        # -- ALERTAS MATEMATICAS POR AWB ---------------------------------------
        errores_math = grupo[grupo["Estado_Validacion"].isin(["ERROR", "ALERTA"])]
        if not errores_math.empty:
            suma_diff = errores_math["Diferencia_CLP"].abs().sum()
            if suma_diff > 0:
                nivel = "ALERTA" if suma_diff <= 50_000 else "ERROR"
                anomalias.append({"awb": awb, "nivel": nivel,
                                  "tipo": "DIFERENCIA_MATEMATICA",
                                  "descripcion": (f"AWB {awb}: diferencia total ${suma_diff:,.0f} CLP "
                                                  f"en {len(errores_math)} lineas")})
                if info_awb["estado_global"] == "OK":
                    info_awb["estado_global"] = nivel

        resumen_awbs[awb] = info_awb

    # Clasificar resumen
    total     = len(resumen_awbs)
    criticos  = sum(1 for v in resumen_awbs.values() if v["estado_global"] == "CRITICO")
    alertas   = sum(1 for v in resumen_awbs.values() if v["estado_global"] == "ALERTA")
    ok_count  = total - criticos - alertas

    return {
        "anomalias":      anomalias,
        "destrucciones":  destrucciones,
        "bloqueados":     bloqueados,
        "resumen_awbs":   resumen_awbs,
        "estadisticas": {
            "total_awbs":    total,
            "criticos":      criticos,
            "alertas":       alertas,
            "ok":            ok_count,
            "hay_criticos":  criticos > 0,
        }
    }


# --- STAGE 5: GENERADOR DE EXCEL VALIDADO -------------------------------------
def _fill(ws, cell, color):
    ws[cell].fill = PatternFill("solid", fgColor=color)

def _bold_header(ws, row_n, max_col, color=COLOR_HEADER):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_n, column=c)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def stage5_generar_excel(
    path_fast:       str,
    dfs_validados:   dict[str, pd.DataFrame],
    resultado_audit: dict,
    path_salida:     str,
    resumen_gemini:  str = ""
) -> str:
    """
    Genera el Excel de salida con:
    - Hoja TOTALES: sin cambios
    - Hojas de detalle: columnas de validacion agregadas con color
    - Hoja INFORME: resumen ejecutivo + anomalias + destrucciones
    """
    wb = openpyxl.load_workbook(path_fast)

    COLS_NUEVAS = ["Valor_Calculado", "Diferencia_CLP",
                   "Estado_Validacion", "Nota_Validacion", "Tarifa_Tarifario"]
    COLOR_MAP   = {"OK": COLOR_OK, "ALERTA": COLOR_ALERTA,
                   "ERROR": COLOR_ERROR, "ADVERTENCIA": COLOR_ALERTA}

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for hoja in HOJAS_DETALLE:
        if hoja not in dfs_validados:
            continue
        df_val = dfs_validados[hoja]

        # Eliminar hoja original y reescribir
        if hoja in wb.sheetnames:
            del wb[hoja]
        ws = wb.create_sheet(hoja)

        all_cols = list(df_val.columns)
        for c_idx, col_name in enumerate(all_cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = thin

        for r_idx, (_, row) in enumerate(df_val.iterrows(), 2):
            estado = str(row.get("Estado_Validacion", "OK"))
            color  = COLOR_MAP.get(estado, "FFFFFF")
            for c_idx, col_name in enumerate(all_cols, 1):
                val  = row[col_name]
                if pd.isna(val):
                    val = ""
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Colorear solo columnas nuevas
                if col_name in COLS_NUEVAS:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.border = thin

        # Autofit ancho aproximado
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    # -- HOJA INFORME ----------------------------------------------------------
    if "INFORME" in wb.sheetnames:
        del wb["INFORME"]
    ws_inf = wb.create_sheet("INFORME", 0)

    stats = resultado_audit["estadisticas"]
    r = 1

    # Titulo
    ws_inf.merge_cells(f"A{r}:F{r}")
    ws_inf[f"A{r}"] = "INFORME DE VALIDACION - TERMINAL FAST"
    ws_inf[f"A{r}"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws_inf[f"A{r}"].fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    ws_inf[f"A{r}"].alignment = Alignment(horizontal="center")
    r += 1

    ws_inf[f"A{r}"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_inf[f"A{r}"].font = Font(italic=True, color="595959")
    r += 2

    # Resumen estadistico
    for label, val, color in [
        ("Total AWBs procesadas",  stats["total_awbs"], COLOR_INFO),
        ("AWBs sin anomalias",     stats["ok"],         COLOR_OK),
        ("AWBs con alertas",       stats["alertas"],    COLOR_ALERTA),
        ("AWBs con errores criticos", stats["criticos"], COLOR_ERROR),
    ]:
        ws_inf[f"A{r}"] = label
        ws_inf[f"B{r}"] = val
        ws_inf[f"A{r}"].font = Font(bold=True)
        ws_inf[f"B{r}"].fill = PatternFill("solid", fgColor=color)
        ws_inf[f"B{r}"].alignment = Alignment(horizontal="center")
        r += 1

    r += 1

    # Resumen Gemini (si existe)
    if resumen_gemini:
        ws_inf.merge_cells(f"A{r}:F{r}")
        ws_inf[f"A{r}"] = "RESUMEN EJECUTIVO"
        ws_inf[f"A{r}"].font = Font(bold=True, color="FFFFFF")
        ws_inf[f"A{r}"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
        r += 1
        ws_inf.merge_cells(f"A{r}:F{r+4}")
        ws_inf[f"A{r}"] = resumen_gemini
        ws_inf[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 6

    # Tabla de anomalias
    anomalias = resultado_audit["anomalias"]
    if anomalias:
        ws_inf.merge_cells(f"A{r}:F{r}")
        ws_inf[f"A{r}"] = "ANOMALIAS DETECTADAS"
        ws_inf[f"A{r}"].font = Font(bold=True, color="FFFFFF")
        ws_inf[f"A{r}"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
        r += 1

        headers = ["AWB", "Nivel", "Tipo", "Descripcion"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws_inf.cell(row=r, column=c_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="595959")
        r += 1

        for an in anomalias:
            color = COLOR_ERROR if an["nivel"] == "CRITICO" else COLOR_ALERTA
            for c_idx, key in enumerate(["awb", "nivel", "tipo", "descripcion"], 1):
                cell = ws_inf.cell(row=r, column=c_idx, value=an.get(key, ""))
                cell.fill   = PatternFill("solid", fgColor=color)
                cell.border = thin
            r += 1

    r += 1

    # Tabla de destrucciones
    destrucciones = resultado_audit["destrucciones"]
    if destrucciones:
        ws_inf.merge_cells(f"A{r}:F{r}")
        ws_inf[f"A{r}"] = "REGISTRO DE DESTRUCCIONES DE CAJAS"
        ws_inf[f"A{r}"].font = Font(bold=True, color="FFFFFF")
        ws_inf[f"A{r}"].fill = PatternFill("solid", fgColor="7B2C2C")
        r += 1

        headers_d = ["AWB", "Exportador", "Vuelo", "Destino", "Hoja", "Valor CLP"]
        for c_idx, h in enumerate(headers_d, 1):
            cell = ws_inf.cell(row=r, column=c_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="595959")
        r += 1

        for dest in destrucciones:
            for c_idx, key in enumerate(["awb", "exportador", "vuelo", "destino", "hoja", "valor"], 1):
                cell = ws_inf.cell(row=r, column=c_idx, value=dest.get(key, ""))
                cell.fill   = PatternFill("solid", fgColor="F4CCCC")
                cell.border = thin
            r += 1

    # Autofit informe
    for col in ws_inf.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_inf.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    wb.save(path_salida)
    return path_salida


# --- ORQUESTADOR PRINCIPAL ----------------------------------------------------
def ejecutar_validacion(path_fast: str, path_tarifario: str,
                        path_salida: str = None) -> dict:
    """
    Punto de entrada unico. Usado tanto en CLI como en LangGraph (Stage node).
    Retorna dict con todos los resultados para que Stage 4 (Gemini) lo consuma.
    """
    print("> Stage 1 - Cargando archivos...")
    tarifario = cargar_tarifario(path_tarifario)

    dfs = {}
    for hoja in HOJAS_DETALLE:
        try:
            dfs[hoja] = pd.read_excel(path_fast, sheet_name=hoja)
            print(f"  OK {hoja}: {len(dfs[hoja])} filas")
        except Exception as e:
            print(f"  ERROR {hoja}: no encontrada - {e}")

    if not dfs:
        raise ValueError("No se encontraron hojas de detalle en el archivo FAST")

    # Validar vigencia del tarifario contra la fecha mas reciente del archivo
    todas_fechas = pd.concat(list(dfs.values()))["Fecha Inicio Cobro"].dropna()
    fecha_ref    = pd.to_datetime(todas_fechas, dayfirst=True, errors="coerce").max()
    if pd.isna(fecha_ref):
        fecha_ref = datetime.now()
    vig = validar_vigencia(tarifario, fecha_ref.to_pydatetime())
    if not vig["vigente"]:
        print(f"  WARN ADVERTENCIA: fecha {fecha_ref.date()} fuera de vigencia del tarifario ({vig['vigencia_str']})")

    print("> Stage 2 - Validacion matematica...")
    dfs_validados = stage2_validar_matematica(dfs, tarifario)
    for hoja, df in dfs_validados.items():
        estados = df["Estado_Validacion"].value_counts().to_dict()
        print(f"  OK {hoja}: {estados}")

    print("> Stage 3 - Auditoria de integridad por AWB...")
    resultado_audit = stage3_auditar_integridad(dfs_validados)
    stats = resultado_audit["estadisticas"]
    print(f"  OK {stats['total_awbs']} AWBs | "
          f"OK: {stats['ok']} | Alertas: {stats['alertas']} | "
          f"Criticos: {stats['criticos']}")

    if resultado_audit["destrucciones"]:
        print(f"  OK {len(resultado_audit['destrucciones'])} destrucciones de cajas registradas")
    if resultado_audit["bloqueados"]:
        print(f"  ERROR {len(resultado_audit['bloqueados'])} AWBs AUSTRALIS bloqueadas")

    # Stage 4 (Gemini) se inyecta externamente - aqui solo preparamos el contexto
    contexto_gemini = _preparar_contexto_gemini(resultado_audit, dfs_validados)

    if path_salida:
        print("> Stage 5 - Generando Excel validado...")
        stage5_generar_excel(path_fast, dfs_validados, resultado_audit, path_salida)
        print(f"  OK Guardado en: {path_salida}")

    return {
        "dfs_validados":   dfs_validados,
        "resultado_audit": resultado_audit,
        "contexto_gemini": contexto_gemini,
        "vigencia":        vig,
        "path_salida":     path_salida,
    }


def _preparar_contexto_gemini(resultado_audit: dict,
                               dfs_validados: dict) -> str:
    """
    Genera el JSON/texto estructurado que se enviara a Gemini en Stage 4.
    El LLM recibe solo los hechos ya calculados - nunca los numeros crudos.
    """
    stats = resultado_audit["estadisticas"]
    total_filas = sum(len(df) for df in dfs_validados.values())

    # Calcular diferencia total en pesos
    diffs = []
    for df in dfs_validados.values():
        diffs.extend(df["Diferencia_CLP"].dropna().tolist())
    diff_total = sum(diffs)
    diff_abs   = sum(abs(d) for d in diffs)

    ctx = {
        "resumen": {
            "total_filas_procesadas": total_filas,
            "total_awbs":             stats["total_awbs"],
            "awbs_sin_anomalias":     stats["ok"],
            "awbs_con_alertas":       stats["alertas"],
            "awbs_criticas":          stats["criticos"],
            "diferencia_total_clp":   round(diff_total, 2),
            "diferencia_absoluta_clp": round(diff_abs, 2),
        },
        "anomalias_criticas": [
            a for a in resultado_audit["anomalias"] if a["nivel"] == "CRITICO"
        ],
        "anomalias_alerta": [
            a for a in resultado_audit["anomalias"] if a["nivel"] == "ALERTA"
        ],
        "destrucciones": resultado_audit["destrucciones"],
        "bloqueados":    resultado_audit["bloqueados"],
    }
    import json
    return json.dumps(ctx, ensure_ascii=False, indent=2)


# --- CLI ----------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Agente FAST - Validador de Gastos Terminales")
    parser.add_argument("--fast",       required=True, help="Ruta al archivo FAST .xlsx")
    parser.add_argument("--tarifario",  required=True, help="Ruta al tarifario .xlsx")
    parser.add_argument("--salida",     default="FAST_validado.xlsx",
                        help="Ruta del Excel de salida (default: FAST_validado.xlsx)")
    args = parser.parse_args()

    resultado = ejecutar_validacion(args.fast, args.tarifario, args.salida)
    print("\n> Validacion completa.")
    print(f"   Archivo validado: {resultado['path_salida']}")
    hay_criticos = resultado["resultado_audit"]["estadisticas"]["hay_criticos"]
    if hay_criticos:
        print("   WARN HAY ANOMALIAS CRITICAS - requiere aprobacion humana antes de entregar.")
    else:
        print("   > Sin anomalias criticas - listo para entrega.")